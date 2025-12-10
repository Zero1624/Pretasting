"""
Flask backend for Pre-Taste feedback system
Stores feedback submissions in an Excel file
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from pathlib import Path

app = Flask(__name__)
CORS(app)

# Path to feedback Excel file
FEEDBACK_FILE = Path(__file__).parent / 'feedback.xlsx'

# Column headers
HEADERS = ['Timestamp', 'Name', 'Topic', 'Message']

def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not FEEDBACK_FILE.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Feedback'
        
        # Add headers
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="B87333", end_color="B87333", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        wb.save(FEEDBACK_FILE)

def add_feedback_to_excel(name, topic, message):
    """Add a feedback entry to the Excel file"""
    wb = openpyxl.load_workbook(FEEDBACK_FILE)
    ws = wb.active
    
    # Get next empty row
    next_row = ws.max_row + 1
    
    # Prepare data
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = [timestamp, name or '(Not provided)', topic or '(Not specified)', message]
    
    # Add data to row
    thin_border = Border(
        left=Side(style='thin', color='E5D5C5'),
        right=Side(style='thin', color='E5D5C5'),
        top=Side(style='thin', color='E5D5C5'),
        bottom=Side(style='thin', color='E5D5C5')
    )
    
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border
        if col == 4:  # Message column
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust row height for wrapped text
    ws.row_dimensions[next_row].height = None
    
    wb.save(FEEDBACK_FILE)

@app.route('/api/feedback', methods=['POST'])
def submit_feedback():
    """
    API endpoint to receive and store feedback
    Expected JSON body:
    {
        "name": "John Doe",
        "topic": "food",
        "message": "Great appetizers!"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Invalid JSON'}), 400
        
        # Extract fields
        name = data.get('name', '').strip()
        topic = data.get('topic', '').strip()
        message = data.get('message', '').strip()
        
        # Validate message (required)
        if not message:
            return jsonify({'error': 'Message is required'}), 400
        
        # Ensure Excel file exists
        initialize_excel()
        
        # Add feedback to Excel
        add_feedback_to_excel(name, topic, message)
        
        return jsonify({
            'success': True,
            'message': 'Feedback received and saved successfully'
        }), 200
        
    except Exception as e:
        print(f'Error processing feedback: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/feedback/list', methods=['GET'])
def list_feedback():
    """
    Retrieve all feedback entries (for admin purposes)
    """
    try:
        if not FEEDBACK_FILE.exists():
            return jsonify({'feedback': []}), 200
        
        wb = openpyxl.load_workbook(FEEDBACK_FILE)
        ws = wb.active
        
        feedback_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                feedback_list.append({
                    'timestamp': row[0],
                    'name': row[1],
                    'topic': row[2],
                    'message': row[3]
                })
        
        return jsonify({'feedback': feedback_list}), 200
        
    except Exception as e:
        print(f'Error retrieving feedback: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'ok'}), 200

if __name__ == '__main__':
    # Initialize Excel file on startup
    initialize_excel()
    
    # Run Flask app
    print('Pre-Taste Feedback Server running on http://localhost:5000')
    print(f'Feedback file: {FEEDBACK_FILE}')
    app.run(debug=True, port=5000)
